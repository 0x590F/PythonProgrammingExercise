import openpyxl

def main():
    wb = openpyxl.load_workbook(r"C:\Users\zhangyiduo\Desktop\固定资产已领用信息20201222.xlsx")
    ws = wb["固定资产已领用信息"]
    nws = wb.create_sheet(index = 0,title = "New sheet")
    list=[]
    for row in ws.iter_rows(min_row = 2, max_row = ws.max_row):
        if float(row[4].value) > 50000.00:
            sub_list=[]
            for i in range(0,10):
                sub_list.append(row[i].value)
            list.append(sub_list)
    print(list)
    for row in list:
        nws.append(row)
    wb.save(r"C:\Users\zhangyiduo\Desktop\固定资产已领用信息20201222.xlsx")

main()
